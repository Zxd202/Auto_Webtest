from openpyxl import load_workbook


#封装一个读取xlsx的类,按行读取
class ParseXLSX():
    def parse_xlsx(file,sheet,col):
        mylist = []
        wb = load_workbook(file)
        cur_sheet = wb.get_sheet_by_name(sheet)
        for i in range(1,cur_sheet.max_row+1):
            value1 = cur_sheet["{}{}".format(col,i)].value
            mylist.append(value1)

        #print(mylist)
        return mylist


#封装一个读取xlsx的类，按列读取



if __name__ == '__main__':
    myfile = 'user.xlsx'
    mysheet = 'Sheet1'
    print(ParseXLSX.parse_xlsx(myfile,mysheet,'C'))



